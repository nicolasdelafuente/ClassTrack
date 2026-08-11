"""Extract ClassTrack seed data from the local Excel (not committed).

Primary roster: sheet DesApp2026 (from row 3)
  B = Legajo (overridden to 40000000 for local tests)
  C = Alumno (full name)
  D = Grupo asignado
  E = Email (Contactos)
  F+ = Presentismo (fecha en fila 1; ASISTENCIA / PARTICIPACION / APROBACION)

Extra sheets only enrich topic / teacher / links.
"""
from __future__ import annotations

import json
import re
from datetime import datetime
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[4]
if not (ROOT / "DesAPP-PPS 2026-c1- Asistencia.xlsx").exists():
    ROOT = Path(__file__).resolve().parent
    while ROOT.parent != ROOT and not (ROOT / "DesAPP-PPS 2026-c1- Asistencia.xlsx").exists():
        ROOT = ROOT.parent

XLSX = ROOT / "DesAPP-PPS 2026-c1- Asistencia.xlsx"
OUT_DIR = Path(__file__).resolve().parents[1] / "data"  # prisma/data
OUT_FULL = OUT_DIR / "from-excel.json"
OUT_DEMO = OUT_DIR / "demo.json"

# Forced legajo for local/realismo tests (user request).
FORCED_LEGAJO = "40000000"


def group_number(value: object) -> int | None:
    if value is None:
        return None
    m = re.search(r"(\d+)", str(value))
    return int(m.group(1)) if m else None


def ensure_group(groups: dict[int, dict], num: int) -> dict:
    if num not in groups:
        groups[num] = {
            "number": num,
            "name": f"Grupo {num}",
            "projectTopic": None,
            "teacherName": None,
            "students": [],
            "links": {},
        }
    return groups[num]


def is_marked_present(value: object) -> bool:
    if value is None:
        return False
    return str(value).strip().lower() == "p"


def is_marked_participated(value: object) -> bool:
    if value is None:
        return False
    token = str(value).strip().lower()
    return token in {"p", "x", "1", "si", "sí", "true"}


def attendance_date_columns(ws) -> list[tuple[int, int, str]]:
    """Return (asistencia_col, participacion_col, YYYY-MM-DD) from row 1/2."""
    cols: list[tuple[int, int, str]] = []
    c = 6
    while c <= ws.max_column:
        raw = ws.cell(1, c).value
        if isinstance(raw, datetime):
            cols.append((c, c + 1, raw.strftime("%Y-%m-%d")))
        c += 3
    return cols


def main() -> None:
    if not XLSX.exists():
        raise SystemExit(f"Excel not found: {XLSX}")

    print("ROOT", ROOT)
    print("XLSX", XLSX)

    wb = openpyxl.load_workbook(XLSX, data_only=True)
    groups: dict[int, dict] = {}

    # --- Primary: DesApp2026 roster + assigned groups ---
    ws = wb["DesApp2026"]
    seen_emails: set[str] = set()
    for row in ws.iter_rows(min_row=3, max_col=5, values_only=True):
        _n, _legajo, alumno, grupo, contactos = (list(row) + [None] * 5)[:5]
        if not alumno:
            continue
        gnum = group_number(grupo)
        if gnum is None:
            print(f"skip (no group): {alumno}")
            continue

        name = str(alumno).strip()
        email = str(contactos).strip() if contactos else None
        if email and "@" not in email:
            email = None

        g = ensure_group(groups, gnum)
        if email:
            key = email.lower()
            if key in seen_emails:
                print(f"skip duplicate email: {email}")
                continue
            seen_emails.add(key)

        g["students"].append(
            {
                "fullName": name,
                "email": email,
                "legajo": FORCED_LEGAJO,
            }
        )

    # --- Presentismo (same sheet, date blocks from col F) ---
    date_cols = attendance_date_columns(ws)
    attendance: list[dict] = []
    for r in range(3, ws.max_row + 1):
        alumno = ws.cell(r, 3).value
        contactos = ws.cell(r, 5).value
        if not alumno or not contactos:
            continue
        email = str(contactos).strip().lower()
        if "@" not in email or email not in seen_emails:
            continue
        for asist_c, part_c, iso_date in date_cols:
            attendance.append(
                {
                    "email": email,
                    "date": iso_date,
                    "present": is_marked_present(ws.cell(r, asist_c).value),
                    "participated": is_marked_participated(
                        ws.cell(r, part_c).value
                    ),
                }
            )

    # --- Enrich: topic + teacher from "DesApp por grupo" ---
    if "DesApp por grupo" in wb.sheetnames:
        ws_g = wb["DesApp por grupo"]
        current: int | None = None
        for row in ws_g.iter_rows(min_row=1, max_col=6, values_only=True):
            a, _b, _c, d, e, _f = (list(row) + [None] * 6)[:6]
            if a and str(a).startswith("Grupo"):
                current = group_number(a)
                if current is None:
                    continue
                g = ensure_group(groups, current)
                topic = d.strip() if isinstance(d, str) and d.strip() else None
                teacher = e.strip() if isinstance(e, str) and e.strip() else None
                if topic:
                    g["projectTopic"] = topic
                if teacher:
                    g["teacherName"] = teacher

    # --- Enrich: links from inicializacion ---
    if "Desapp - inicializacion" in wb.sheetnames:
        ws_i = wb["Desapp - inicializacion"]
        for row in ws_i.iter_rows(min_row=2, max_col=4, values_only=True):
            gname, github, trello, folder = (list(row) + [None] * 4)[:4]
            num = group_number(gname)
            if num is None:
                continue
            g = ensure_group(groups, num)
            g["links"] = {
                "githubUrl": str(github).strip() if github else None,
                "trelloUrl": str(trello).strip() if trello else None,
                "driveUrl": None,
                "folderName": str(folder).strip() if folder else None,
            }

    # --- Enrich: teacher from DESAPP FINAL ---
    if "DESAPP FINAL" in wb.sheetnames:
        ws_f = wb["DESAPP FINAL"]
        for row in ws_f.iter_rows(min_row=2, max_col=3, values_only=True):
            g, _mid, teacher = (list(row) + [None] * 3)[:3]
            num = group_number(g)
            if num is None or num not in groups or not teacher:
                continue
            groups[num]["teacherName"] = str(teacher).strip().title()

    payload = {
        "course": {
            "name": "Desarrollo de Aplicaciones 2026-c1",
            "code": "2026-c1",
            "isCurrent": True,
        },
        "groups": [groups[n] for n in sorted(groups.keys())],
        "attendance": attendance,
    }

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    OUT_FULL.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    # Anonymized demo for the public repo (no real emails/names)
    email_to_demo: dict[str, str] = {}
    demo = {
        "course": payload["course"],
        "groups": [],
        "attendance": [],
    }
    for g in payload["groups"]:
        demo_students = []
        for i, s in enumerate(g["students"], start=1):
            demo_email = f"alumno{g['number']:02d}{i:02d}@demo.classtrack.local"
            if s.get("email"):
                email_to_demo[str(s["email"]).lower()] = demo_email
            demo_students.append(
                {
                    "fullName": f"Alumno {g['number']}-{i:02d}",
                    "email": demo_email,
                    "legajo": f"2026{g['number']:02d}{i:02d}",
                }
            )
        demo["groups"].append(
            {
                "number": g["number"],
                "name": g["name"],
                "projectTopic": g.get("projectTopic"),
                "teacherName": g.get("teacherName"),
                "students": demo_students,
                "links": g.get("links") or {},
            }
        )

    for rec in attendance:
        demo_email = email_to_demo.get(rec["email"])
        if not demo_email:
            continue
        demo["attendance"].append(
            {
                "email": demo_email,
                "date": rec["date"],
                "present": rec["present"],
                "participated": rec["participated"],
            }
        )

    OUT_DEMO.write_text(
        json.dumps(demo, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    present_marks = sum(1 for a in attendance if a["present"])
    total = sum(len(g["students"]) for g in payload["groups"])
    print(
        f"groups={len(payload['groups'])} students={total} "
        f"legajo={FORCED_LEGAJO} attendance={len(attendance)} "
        f"present={present_marks} dates={len(date_cols)}"
    )
    for g in payload["groups"]:
        print(
            f"G{g['number']}: {len(g['students'])} students | "
            f"{g.get('projectTopic')} | {g.get('teacherName')}"
        )
    print(f"wrote {OUT_FULL}")
    print(f"wrote {OUT_DEMO} (anonymized)")


if __name__ == "__main__":
    main()
